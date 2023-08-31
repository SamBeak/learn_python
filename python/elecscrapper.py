from requests import get
from bs4 import BeautifulSoup
from extractors.wwr import extract_wwr_jobs
from selenium import webdriver

import json # json data
import openpyxl # excel data

filename = "business_listings2.xlsx"
wb = openpyxl.Workbook()
sheet = wb.active


business_name_list = []
with open('./python/business_name_list.json', 'r', encoding='utf-8') as f:
    business_name_list = json.load(f)

browser = webdriver.Chrome()

base_url = "https://www.ecic.kr/ecic/ad/ad0101.do?menuCd=6047&currentPageNo=1&license=&gubun=&sido=&hoeyn=&searchSido=&searchSigungu=&searchGubun=company&searchText="

for item in business_name_list:
    for i in range(len(business_name_list[item])):
        search_term = business_name_list[item][i]
        browser.get(f"{base_url}{search_term}")     # selenium 을 사용해서 scraping을 해오겠다.

        results = []
        soup = BeautifulSoup(browser.page_source, "html.parser")    # 셀레니움이 가져온 페이지 소스
        business_list = soup.find("table", class_="txtC")


        tds = business_list.find_all('td') # recursive=False 는 자식만 가져오겠다는 뜻

        if len(tds) != 1:
            reg_num = tds[0] # reg_num 은 tds의 첫번째 요소
            business_name = tds[1] # business_name 은 tds의 두번째 요소
            boss_name = tds[2] # boss_name 은 tds의 세번째 요소
            location = tds[3] # location 은 tds의 네번째 요소
            anchor = tds[1].find("a")
            link = anchor["href"]
            # link = "javascript:js_detailAction('11000095', '32')" 일 때, '11000095', '32' 를 추출해야함
            license = link.split("'")[1]
            sido = link.split("'")[3]

            business_data = {
                'regnum' : reg_num.string,
                'businessname' : business_name.string,
                'bossname' : boss_name.string,
                'location' : location.string,
                'license' : license,
                'sido' : sido
            }
            results.append(business_data)
            sheet.cell(row=i+1, column=1).value = reg_num.string
            sheet.cell(row=i+1, column=2).value = business_name.string
            sheet.cell(row=i+1, column=3).value = boss_name.string
            sheet.cell(row=i+1, column=4).value = location.string
            sheet.cell(row=i+1, column=5).value = license
            sheet.cell(row=i+1, column=6).value = sido
wb.save(filename)