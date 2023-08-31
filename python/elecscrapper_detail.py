from requests import get
from bs4 import BeautifulSoup
from extractors.wwr import extract_wwr_jobs
from selenium import webdriver

import json # json data
import openpyxl # excel data

filename = "business_listings_detail.xlsx"
wb = openpyxl.Workbook()
sheet = wb.active


business_detail_list = []
with open('./python/business_name_list_detail.json', 'r', encoding='utf-8') as f:
    business_detail_list = json.load(f)


browser = webdriver.Chrome()

base_url = "https://www.ecic.kr/ecic/ad/ad0101D.do?menuCd=6047&currentPageNo=1&license="
add_url = "&sido="
add_url2 = "&searchText="

for item in business_detail_list:
    for i in range(len(business_detail_list[item])):
        search_term = business_detail_list[item][i]["name"]
        license = business_detail_list[item][i]["license"]
        sido = business_detail_list[item][i]["sido"]
        print(f"{base_url}{license}{add_url}{sido}{add_url2}{search_term}")
        browser.get(f"{base_url}{license}{add_url}{sido}{add_url2}{search_term}")     # selenium 을 사용해서 scraping을 해오겠다.

        results = []
        soup = BeautifulSoup(browser.page_source, "html.parser")
        business_list = soup.find("caption", string="회원사정보").parent
        business_list2 = soup.find("caption", string="시공능력평가").parent
        trs = business_list.find_all('tr')
        trs2 = business_list2.find_all('tr')

        for tr in trs:
            th = tr.find('th')
            if th.string == "등록일":
                reg_date = tr.find('td')
            elif th.string == "상호":
                business_name = tr.find('td')
        for tr in trs2:
            th = tr.find('th')
            if th.string == "시공능력평가액":
                evaluation = tr.find('td')
            elif th.string == "지역순위":
                # td안에 strong태그를 제외한 내용 가져오기
                local_grade = tr.find('td').find('strong').string
                local_scale = ''.join(tr.find('td').findAll(text=True, recursive=False))
                local = local_grade + local_scale
            elif th.string == "전국순위":
                global_grade = tr.find('td').find('strong').string
                global_scale = ''.join(tr.find('td').findAll(text=True, recursive=False))
                global_ = global_grade + global_scale
                global_grade_num = ""
                for args in global_grade:
                    if args.isdigit():
                        global_grade_num += args
        business_data = {
            'regdate': reg_date.string,
            'businessname': business_name.string,
            'evaluation': evaluation.string,
            'local': local,
            'global': global_,
            'global_grade_num': global_grade_num
        }
        results.append(business_data)
        sheet.cell(row=i+1, column=1).value = business_data['regdate']
        sheet.cell(row=i+1, column=2).value = business_data['businessname']
        sheet.cell(row=i+1, column=3).value = business_data['evaluation']
        sheet.cell(row=i+1, column=4).value = business_data['local']
        sheet.cell(row=i+1, column=5).value = business_data['global']
        sheet.cell(row=i+1, column=6).value = business_data['global_grade_num']
wb.save(filename)